from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy, reverse
from django.db.models import Q
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse, QueryDict, HttpResponse
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .models import Supervisor, Conductor, EntregaLibreta
from .forms import SupervisorForm, ConductorForm, EntregaLibretaForm
import calendar
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.contrib.auth.mixins import UserPassesTestMixin
import logging # Import the logging module
from django.contrib.auth.models import User
from django.conf import settings
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

logger = logging.getLogger(__name__) # Get a logger instance

class HomeView(View):
    def get(self, request):
        if request.user.is_authenticated:
            try:
                supervisor = request.user.supervisor_profile
                supervisores = [supervisor]
            except Supervisor.DoesNotExist:
                if request.user.is_staff:
                    supervisores = Supervisor.objects.all()
                else:
                    supervisores = []
        else:
            supervisores = []
            
        mes_actual = timezone.now().month
        anio_actual = timezone.now().year
        
        return render(request, 'libretas/home.html', {
            'supervisores': supervisores,
            'mes_actual': mes_actual,
            'anio_actual': anio_actual,
        })

class SupervisorListView(LoginRequiredMixin, ListView):
    model = Supervisor
    template_name = 'libretas/supervisor_list.html'
    context_object_name = 'supervisores'

    def get_queryset(self):
        if self.request.user.is_staff:
            return Supervisor.objects.all()
        try:
            return [self.request.user.supervisor_profile]
        except Supervisor.DoesNotExist:
            return Supervisor.objects.none()

class StaffRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_staff

    def handle_no_permission(self):
        messages.error(self.request, 'No tienes permisos para realizar esta acción. Solo los administradores pueden realizar esta operación.')
        return redirect('home')

class SupervisorCreateView(StaffRequiredMixin, CreateView):
    model = Supervisor
    form_class = SupervisorForm
    template_name = 'libretas/supervisor_form.html'
    success_url = reverse_lazy('supervisor-list')

    def form_valid(self, form):
        messages.success(self.request, 'Supervisor creado exitosamente.')
        return super().form_valid(form)

class SupervisorUpdateView(LoginRequiredMixin, UpdateView):
    model = Supervisor
    form_class = SupervisorForm
    template_name = 'libretas/supervisor_form.html'
    success_url = reverse_lazy('supervisor-list')

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        if not self.request.user.is_staff and obj != self.request.user.supervisor_profile:
            messages.error(self.request, 'No tienes permisos para editar este supervisor.')
            raise PermissionDenied
        return obj

    def form_valid(self, form):
        messages.success(self.request, 'Supervisor actualizado exitosamente.')
        return super().form_valid(form)

class SupervisorDeleteView(StaffRequiredMixin, DeleteView):
    model = Supervisor
    template_name = 'libretas/supervisor_confirm_delete.html'
    success_url = reverse_lazy('supervisor-list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Supervisor eliminado exitosamente.')
        return super().delete(request, *args, **kwargs)

class ConductorListView(LoginRequiredMixin, ListView):
    model = Conductor
    template_name = 'libretas/conductor_list.html'
    context_object_name = 'conductores'
    
    def get_queryset(self):
        queryset = super().get_queryset().order_by('nombre')  # Ordenar alfabéticamente por nombre
        supervisor_id = self.request.GET.get('supervisor')
        
        if self.request.user.is_staff:
            if supervisor_id:
                queryset = queryset.filter(supervisor_id=supervisor_id)
        else:
            try:
                supervisor = self.request.user.supervisor_profile
                queryset = queryset.filter(supervisor=supervisor)
            except Supervisor.DoesNotExist:
                queryset = Conductor.objects.none()
                
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if not self.request.user.is_staff:
            try:
                self.request.user.supervisor_profile
            except Supervisor.DoesNotExist:
                messages.error(self.request, 'No tienes un perfil de supervisor asociado. Por favor, contacta al administrador.')
                return redirect('home')
        # Añadir supervisores y selección actual al contexto para el filtro
        supervisor_id = self.request.GET.get('supervisor')
        context['supervisor_seleccionado'] = supervisor_id
        if self.request.user.is_staff:
            context['supervisores'] = Supervisor.objects.all().order_by('nombre')
        else:
            try:
                supervisor = self.request.user.supervisor_profile
                context['supervisores'] = Supervisor.objects.filter(id=supervisor.id)
            except Supervisor.DoesNotExist:
                context['supervisores'] = Supervisor.objects.none()
        return context

class ConductorCreateView(LoginRequiredMixin, CreateView):
    model = Conductor
    form_class = ConductorForm
    template_name = 'libretas/conductor_form.html'
    success_url = reverse_lazy('conductor-list')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        if not self.request.user.is_staff:
            try:
                supervisor = self.request.user.supervisor_profile
                form.fields['supervisor'].queryset = Supervisor.objects.filter(id=supervisor.id)
                form.fields['supervisor'].initial = supervisor
                form.fields['supervisor'].widget.attrs['readonly'] = True
            except Supervisor.DoesNotExist:
                form.fields['supervisor'].queryset = Supervisor.objects.none()
                form.fields['supervisor'].widget.attrs['disabled'] = True
        return form

    def form_valid(self, form):
        if not self.request.user.is_staff:
            try:
                supervisor = self.request.user.supervisor_profile
                form.instance.supervisor = supervisor
            except Supervisor.DoesNotExist:
                messages.error(self.request, 'No tienes un perfil de supervisor asociado. Por favor, contacta al administrador.')
                return redirect('home')
        messages.success(self.request, 'Conductor creado exitosamente.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if not self.request.user.is_staff:
            try:
                self.request.user.supervisor_profile
            except Supervisor.DoesNotExist:
                messages.error(self.request, 'No tienes un perfil de supervisor asociado. Por favor, contacta al administrador.')
                # Opcional: podrías agregar una bandera para deshabilitar el formulario
                context['no_supervisor'] = True
        return context

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_staff:
            try:
                request.user.supervisor_profile
            except Supervisor.DoesNotExist:
                messages.error(request, 'No tienes un perfil de supervisor asociado. Por favor, contacta al administrador.')
                return redirect('home')
        return super().dispatch(request, *args, **kwargs)

class ConductorUpdateView(LoginRequiredMixin, UpdateView):
    model = Conductor
    form_class = ConductorForm
    template_name = 'libretas/conductor_form.html'
    success_url = reverse_lazy('conductor-list')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        if not self.request.user.is_staff:
            form.fields['supervisor'].queryset = Supervisor.objects.filter(usuario=self.request.user)
            form.fields['supervisor'].widget.attrs['readonly'] = True
        return form

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        if not self.request.user.is_staff:
            try:
                if obj.supervisor != self.request.user.supervisor_profile:
                    messages.error(self.request, 'No tienes permisos para editar este conductor.')
                    raise PermissionDenied
            except Supervisor.DoesNotExist:
                messages.error(self.request, 'No tienes un perfil de supervisor asociado.')
                raise PermissionDenied
        return obj

    def form_valid(self, form):
        messages.success(self.request, 'Conductor actualizado exitosamente.')
        return super().form_valid(form)

class ConductorDeleteView(LoginRequiredMixin, DeleteView):
    model = Conductor
    template_name = 'libretas/conductor_confirm_delete.html'
    success_url = reverse_lazy('conductor-list')

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        if not self.request.user.is_staff:
            try:
                if obj.supervisor != self.request.user.supervisor_profile:
                    messages.error(self.request, 'No tienes permisos para eliminar este conductor.')
                    raise PermissionDenied
            except Supervisor.DoesNotExist:
                messages.error(self.request, 'No tienes un perfil de supervisor asociado.')
                raise PermissionDenied
        return obj

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Conductor eliminado exitosamente.')
        return super().delete(request, *args, **kwargs)

class SeguimientoLibretasView(LoginRequiredMixin, View):
    template_name = 'libretas/seguimiento_libretas.html'
    
    def get(self, request):
        # Obtener y validar mes y año
        try:
            mes = int(request.GET.get('mes', timezone.now().month))
            anio = int(request.GET.get('anio', timezone.now().year))
        except (ValueError, TypeError):
            mes = timezone.now().month
            anio = timezone.now().year
            
        supervisor_id = request.GET.get('supervisor')
        estado_filtro = request.GET.get('estado') or ''
        # Paginación: tamaño por página configurable
        try:
            per_page = int(request.GET.get('per_page', 17))
        except (ValueError, TypeError):
            per_page = 17
        
        if request.user.is_staff:
            supervisores = Supervisor.objects.all()
            if supervisor_id and supervisor_id != 'None':
                conductores = Conductor.objects.filter(supervisor_id=supervisor_id).order_by('nombre')
            else:
                conductores = Conductor.objects.all().order_by('nombre')
        else:
            try:
                supervisor = request.user.supervisor_profile
                supervisores = [supervisor]
                conductores = Conductor.objects.filter(supervisor=supervisor).order_by('nombre')
            except Supervisor.DoesNotExist:
                supervisores = []
                conductores = Conductor.objects.none()
                messages.error(request, 'No tienes un perfil de supervisor asociado.')
                return redirect('home')
            
        # Determinar el número de semanas en el mes
        ultimo_dia = calendar.monthrange(anio, mes)[1]
        num_semanas = (ultimo_dia // 7) + (1 if ultimo_dia % 7 > 0 else 0)
        if num_semanas > 5:
            num_semanas = 5
            
        # Generar datos para la tabla
        datos_tabla = []
        for conductor in conductores:
            fila = {
                'conductor': conductor,
                'semanas': []
            }
            
            for semana in range(1, num_semanas + 1):
                try:
                    entrega = EntregaLibreta.objects.get(
                        conductor=conductor,
                        mes=mes,
                        anio=anio,
                        semana=semana
                    )
                except EntregaLibreta.DoesNotExist:
                    entrega = EntregaLibreta(
                        conductor=conductor,
                        mes=mes,
                        anio=anio,
                        semana=semana,
                        estado='PENDIENTE'
                    )

                if getattr(settings, 'VERBOSE_SEGUIMIENTO', False):
                    logger.debug("Conductor: %s, Semana: %s, Estado recuperado: %s", conductor.nombre, semana, entrega.estado)

                fila['semanas'].append({
                    'semana': semana,
                    'entrega': entrega
                })
            
            # Aplicar filtro por estado (si se solicitó) sobre las semanas del conductor
            if estado_filtro in {'PENDIENTE', 'ENTREGADO', 'NO_ENTREGADO', 'DESVINCULADO', 'LICENCIA_MEDICA'}:
                tiene_estado = any(s['entrega'].estado == estado_filtro for s in fila['semanas'])
                if not tiene_estado:
                    continue

            datos_tabla.append(fila)

        # Construir paginación
        paginator = Paginator(datos_tabla, per_page)
        page_number = request.GET.get('page')
        try:
            page_obj = paginator.get_page(page_number)
        except (PageNotAnInteger, EmptyPage):
            page_obj = paginator.get_page(1)

        # Preservar filtros en links de paginación
        querydict = request.GET.copy()
        if 'page' in querydict:
            querydict.pop('page')
        querystring = querydict.urlencode()
            
        context = {
            'supervisores': supervisores,
            'supervisor_seleccionado': supervisor_id,
            'mes': mes,
            'anio': anio,
            'datos_tabla': datos_tabla,  # lista completa por si se necesita
            'page_obj': page_obj,        # lista paginada para render
            'is_paginated': paginator.num_pages > 1,
            'paginator': paginator,
            'num_semanas': range(1, num_semanas + 1),
            'meses': [(i, calendar.month_name[i]) for i in range(1, 13)],
            'anios': range(timezone.now().year - 2, timezone.now().year + 3),
            'puede_editar': request.user.is_staff or hasattr(request.user, 'supervisor_profile'),
            'estado_filtro': estado_filtro,
            'per_page': per_page,
            'querystring': querystring,
            'per_page_choices': [10, 25, 50, 100],
            'start_index': page_obj.start_index(),
        }
        
        return render(request, self.template_name, context)

class ActualizarEstadoLibretaView(LoginRequiredMixin, View):
    def post(self, request):
        if getattr(settings, 'VERBOSE_SEGUIMIENTO', False):
            logger.debug("Datos POST recibidos: %s", request.POST)
        
        try:
            # Obtener datos del formulario
            conductor_id = request.POST.get('conductor_id')
            semana = request.POST.get('semana')
            mes = request.POST.get('mes')
            anio = request.POST.get('anio')
            estado = request.POST.get('estado')

            # Log de los datos recibidos
            if getattr(settings, 'VERBOSE_SEGUIMIENTO', False):
                logger.debug("Datos recibidos - conductor_id: %s, semana: %s, mes: %s, anio: %s, estado: %s", conductor_id, semana, mes, anio, estado)

            # Validar datos
            if not conductor_id:
                logger.error("Falta conductor_id")
                return JsonResponse({
                    'success': False,
                    'message': 'Falta el ID del conductor'
                })
            if not semana:
                logger.error("Falta semana")
                return JsonResponse({
                    'success': False,
                    'message': 'Falta la semana'
                })
            if not mes:
                logger.error("Falta mes")
                return JsonResponse({
                    'success': False,
                    'message': 'Falta el mes'
                })
            if not anio:
                logger.error("Falta anio")
                return JsonResponse({
                    'success': False,
                    'message': 'Falta el año'
                })
            if not estado:
                logger.error("Falta estado")
                return JsonResponse({
                    'success': False,
                    'message': 'Falta el estado'
                })

            try:
                conductor = Conductor.objects.get(id=conductor_id)
            except Conductor.DoesNotExist:
                logger.error("Conductor no encontrado: %s", conductor_id)
                return JsonResponse({'success': False, 'message': 'Conductor no encontrado'})

            # Verificar permisos
            if request.user.is_staff:
                tiene_permiso = True
            else:
                try:
                    supervisor_profile = request.user.supervisor_profile
                except Supervisor.DoesNotExist:
                    logger.error("Usuario %s no tiene perfil de supervisor", request.user.username)
                    return JsonResponse({'success': False, 'message': 'No tienes un perfil de supervisor asociado'})
                if conductor.supervisor != supervisor_profile:
                    logger.warning("Permiso denegado para actualizar conductor %s por usuario %s", conductor_id, request.user.username)
                    return JsonResponse({'success': False, 'message': 'No tienes permiso para actualizar este conductor'})

            # Actualizar o crear el estado
            try:
                entrega, created = EntregaLibreta.objects.get_or_create(
                    conductor=conductor,
                    mes=mes,
                    anio=anio,
                    semana=semana,
                    defaults={'estado': estado}
                )

                if not created:
                    entrega.estado = estado
                    entrega.save()

                if getattr(settings, 'VERBOSE_SEGUIMIENTO', False):
                    logger.debug("Estado actualizado exitosamente para conductor %s, semana %s", conductor_id, semana)
                return JsonResponse({
                    'success': True,
                    'message': 'Estado actualizado exitosamente'
                })

            except Exception as e:
                logger.error("Error al actualizar estado: %s", str(e))
                return JsonResponse({
                    'success': False,
                    'message': f'Error al actualizar el estado: {str(e)}'
                })

        except ValueError as e:
            logger.error("Error al actualizar estado: %s", str(e))
            return JsonResponse({'success': False, 'message': 'Datos inválidos'})
        except Exception:
            logger.exception("Error general en la actualización:")
            return JsonResponse({'success': False, 'message': 'Error inesperado'})

class DashboardView(LoginRequiredMixin, View):
    template_name = 'libretas/dashboard.html'

    def get(self, request):
        import calendar
        from datetime import datetime
        hoy = timezone.now()
        # Buscar el mes y año más reciente con datos si no hay filtro
        mes_get = request.GET.get('mes')
        anio_get = request.GET.get('anio')
        if not mes_get or not anio_get:
            # Buscar la entrega más reciente
            ultima_entrega = EntregaLibreta.objects.order_by('-anio', '-mes').first()
            if ultima_entrega:
                mes_actual = ultima_entrega.mes
                anio_actual = ultima_entrega.anio
            else:
                mes_actual = hoy.month
                anio_actual = hoy.year
        else:
            mes_actual = int(mes_get)
            anio_actual = int(anio_get)

        supervisores_data = []
        supervisores = Supervisor.objects.all()
        for supervisor in supervisores:
            libretas_entregadas = EntregaLibreta.objects.filter(
                conductor__supervisor=supervisor,
                estado='ENTREGADO',
                mes=mes_actual,
                anio=anio_actual
            ).count()
            total_libretas = EntregaLibreta.objects.filter(
                conductor__supervisor=supervisor,
                mes=mes_actual,
                anio=anio_actual
            ).count()
            porcentaje = (libretas_entregadas / total_libretas * 100) if total_libretas > 0 else 0
            supervisores_data.append({
                'nombre': supervisor.nombre,
                'entregadas': libretas_entregadas,
                'total': total_libretas,
                'porcentaje': round(porcentaje, 1)
            })
        supervisores_data.sort(key=lambda x: x['porcentaje'], reverse=True)
        meses = [(i, calendar.month_name[i]) for i in range(1, 13)]
        anios = range(hoy.year - 2, hoy.year + 3)
        context = {
            'supervisores_data': supervisores_data,
            'mes_actual': mes_actual,
            'anio_actual': anio_actual,
            'meses': meses,
            'anios': anios,
        }
        return render(request, self.template_name, context)

class ExportSeguimientoCSVView(LoginRequiredMixin, View):
    def get(self, request):
        try:
            mes = int(request.GET.get('mes', timezone.now().month))
            anio = int(request.GET.get('anio', timezone.now().year))
        except (ValueError, TypeError):
            mes = timezone.now().month
            anio = timezone.now().year
        supervisor_id = request.GET.get('supervisor')
        estado_filtro = request.GET.get('estado') or ''

        if request.user.is_staff:
            if supervisor_id and supervisor_id != 'None':
                conductores = Conductor.objects.filter(supervisor_id=supervisor_id).order_by('nombre')
            else:
                conductores = Conductor.objects.all().order_by('nombre')
        else:
            try:
                supervisor = request.user.supervisor_profile
                conductores = Conductor.objects.filter(supervisor=supervisor).order_by('nombre')
            except Supervisor.DoesNotExist:
                return HttpResponse('No autorizado', status=403)

        ultimo_dia = calendar.monthrange(anio, mes)[1]
        num_semanas = (ultimo_dia // 7) + (1 if ultimo_dia % 7 > 0 else 0)
        if num_semanas > 5:
            num_semanas = 5

        response = HttpResponse(content_type='text/csv')
        filename = f"seguimiento_{anio}_{mes}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)

        headers = ['#', 'Conductor', 'Supervisor'] + [f'Semana {i}' for i in range(1, num_semanas + 1)]
        writer.writerow(headers)

        row_index = 1
        for conductor in conductores:
            semanas_estados = []
            incluir = (estado_filtro == '')
            for semana in range(1, num_semanas + 1):
                try:
                    entrega = EntregaLibreta.objects.get(
                        conductor=conductor,
                        mes=mes,
                        anio=anio,
                        semana=semana
                    )
                except EntregaLibreta.DoesNotExist:
                    entrega = EntregaLibreta(
                        conductor=conductor,
                        mes=mes,
                        anio=anio,
                        semana=semana,
                        estado='PENDIENTE'
                    )
                semanas_estados.append(entrega.get_estado_display())
                if estado_filtro in {'PENDIENTE', 'ENTREGADO', 'NO_ENTREGADO', 'DESVINCULADO', 'LICENCIA_MEDICA'} and entrega.estado == estado_filtro:
                    incluir = True
            if incluir:
                writer.writerow([row_index, conductor.nombre, conductor.supervisor.nombre] + semanas_estados)
                row_index += 1

        return response

# Opcional: Placeholder para Excel (requiere openpyxl o xlsxwriter)
class ExportSeguimientoExcelView(LoginRequiredMixin, View):
    def get(self, request):
        try:
            mes = int(request.GET.get('mes', timezone.now().month))
            anio = int(request.GET.get('anio', timezone.now().year))
        except (ValueError, TypeError):
            mes = timezone.now().month
            anio = timezone.now().year
        supervisor_id = request.GET.get('supervisor')
        estado_filtro = request.GET.get('estado') or ''

        if request.user.is_staff:
            if supervisor_id and supervisor_id != 'None':
                conductores = Conductor.objects.filter(supervisor_id=supervisor_id).order_by('nombre')
            else:
                conductores = Conductor.objects.all().order_by('nombre')
        else:
            try:
                supervisor = request.user.supervisor_profile
                conductores = Conductor.objects.filter(supervisor=supervisor).order_by('nombre')
            except Supervisor.DoesNotExist:
                return HttpResponse('No autorizado', status=403)

        ultimo_dia = calendar.monthrange(anio, mes)[1]
        num_semanas = (ultimo_dia // 7) + (1 if ultimo_dia % 7 > 0 else 0)
        if num_semanas > 5:
            num_semanas = 5

        wb = Workbook()
        ws = wb.active
        ws.title = f"Seguimiento {anio}-{mes}"

        headers = ['#', 'Conductor', 'Supervisor'] + [f'Semana {i}' for i in range(1, num_semanas + 1)]
        ws.append(headers)

        row_index = 1
        for conductor in conductores:
            semanas_estados = []
            incluir = (estado_filtro == '')
            for semana in range(1, num_semanas + 1):
                try:
                    entrega = EntregaLibreta.objects.get(
                        conductor=conductor,
                        mes=mes,
                        anio=anio,
                        semana=semana
                    )
                except EntregaLibreta.DoesNotExist:
                    entrega = EntregaLibreta(
                        conductor=conductor,
                        mes=mes,
                        anio=anio,
                        semana=semana,
                        estado='PENDIENTE'
                    )
                semanas_estados.append(entrega.get_estado_display())
                if estado_filtro in {'PENDIENTE', 'ENTREGADO', 'NO_ENTREGADO', 'DESVINCULADO', 'LICENCIA_MEDICA'} and entrega.estado == estado_filtro:
                    incluir = True
            if incluir:
                ws.append([row_index, conductor.nombre, conductor.supervisor.nombre] + semanas_estados)
                row_index += 1

        # Ajuste básico de ancho de columnas
        for col_idx, _ in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18 if col_idx > 3 else 25
        ws.column_dimensions['A'].width = 6

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"seguimiento_{anio}_{mes}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
